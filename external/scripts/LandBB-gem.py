#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import json
from openpyxl import load_workbook

from sqlalchemy import create_engine, text

SQLALCHEMY_DATABASE_URI = os.environ.get('DATABASE_URL') or \
    'postgresql://' + os.environ.get('PGUSER') + ':' + os.environ.get('PGPASSWORD') + '@' + os.environ.get('PGHOST') + ':' + os.environ.get('PGPORT') + '/' + os.environ.get('PGDATABASE')
sql_insert = text('INSERT INTO externaldata.all_data (datasource_id, all_data) VALUES (:source_id, :json)')
sql_update_mod = text('UPDATE externaldata.datasource SET sourcedate=(:sourcedate)::timestamp WHERE id=:source_id')

try:
    engine = create_engine(SQLALCHEMY_DATABASE_URI)
except:
    print("Unable to connect to the database.")


def parse_file():
    xlsx = load_workbook(filename = '../data/brandenburg/GemVerz.xlsx')
    # print (xlsx.sheetnames)
    sheet = xlsx.active
    last_mod = xlsx.properties.modified
    engine.execute(sql_update_mod, {
        'source_id': 6,
        'sourcedate': last_mod
    })
    rownum = 3 # Starting data of row is 3
    while sheet['A'+str(rownum)].value is not None:
        ars = sheet['B'+str(rownum)].value # Amtlichen Regionalschlüssel
        ags = sheet['G'+str(rownum)].value # Amtlicher Gemeindeschlüssel
        name = sheet['I'+str(rownum)].value
        sorbisch = sheet['J'+str(rownum)].value
        typ = sheet['C'+str(rownum)].value
        json_data = {
            'ars': ars,
            'ags': ags,
            'name': name,
            'sorbisch': sorbisch,
            'status': typ
        }

        engine.execute(sql_insert, {
            'source_id': 6,
            'json': json.dumps(json_data)
        })
        rownum += 1

parse_file()
